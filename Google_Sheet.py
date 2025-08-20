from openpyxl import Workbook, load_workbook
import os

class ExcelHandler:
    def __init__(self, filename="info.xlsx"):
        self.filename = filename

        
        if os.path.exists(self.filename):
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Students Data"
            
            self.ws.append(["ID", "NAME", "MARK"])
            self.wb.save(self.filename)

    def insert(self, row):
        """নতুন ডাটা ইনসার্ট করা"""
        self.ws.append(row)
        self.wb.save(self.filename)

    def read_all(self):
        """সব ডাটা পড়া"""
        data = []
        for row in self.ws.iter_rows(values_only=True):
            data.append(row)
        return data


# ব্যবহার
xls = ExcelHandler()
xls.insert([101, "Alex", 100])
xls.insert([102, "Rafi", 85])
xls.insert([103, "Yeasin", 90])

print(xls.read_all())
    
     
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    